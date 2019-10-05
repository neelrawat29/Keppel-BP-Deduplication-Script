import sys
import numpy as np
import multiprocessing as mp

from openpyxl import workbook, load_workbook

from dedupe_range import DedupeRange


def process_range(start_row, end_row, uen, search_term, full_name, city_zip, source=None):
    dedupe_range = DedupeRange(
        len(uen), start_row, end_row, source=source)
    dedupe_range.process_UEN(uen)
    dedupe_range.process(search_term)
    dedupe_range.process(full_name)
    dedupe_range.process(city_zip, weight=0.5, weight_exact_match=0.5)

    return dedupe_range.score


def dedupe_file(file_path, ignore_same_source=False):

    print("Opening workbook %s" % file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    def iter_col(column_no):
        '''Helper function to iterate through a column of the excel workbook'''
        return [str(cell[0]).upper() for cell in ws.iter_rows(min_row=2, min_col=column_no, max_col=column_no, values_only=True)]

    ids = iter_col(1)
    row_count = len(ids)

    source = iter_col(2) if ignore_same_source else None

    n_proc = mp.cpu_count() - 1
    chunk_size = round(row_count/n_proc + .5)

    def get_row_range():
        '''Partitions the ranges into roughly equal chunks for parallel processing'''
        target_elements = round(row_count ** 2 / n_proc / 2)
        row_range = []
        row_start = 0
        row_end = 0

        for n in range(n_proc):
            elements_current = 0
            while elements_current < target_elements and row_end < row_count:
                elements_current += row_end
                row_end += 1

            row_range.append((row_start, row_end))
            row_start = row_end
            
        return row_range

    row_range = get_row_range()

    with mp.Pool(processes=n_proc) as pool:
        proc_results = [pool.apply_async(process_range, args=(
            *r, iter_col(3), iter_col(4), iter_col(5), iter_col(6), source)) for r in row_range]
        result = [r.get() for r in proc_results]

    score = np.concatenate(result)

    score += score.T

    total_score = sum(np.max(score, axis=0))
    print('Total score  is {} for {} rows (average {})'.format(
        total_score, len(ids), total_score/len(ids)))

    score_column = ws.max_column + 1

    ws.cell(1, score_column).value = 'Similarity Score'

    for i, max_score in enumerate(np.max(score, axis=0)):
        ws.cell(i+2, score_column).value = max_score

    wb.save(filename='output.xlsx')


if __name__ == "__main__":
    file_path = sys.argv[1]

    ignore_same_source = False if len(sys.argv) <= 2 else sys.argv[2]

    dedupe_file(file_path, ignore_same_source)
