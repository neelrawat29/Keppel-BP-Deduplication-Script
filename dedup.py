import sys
import re
from functools import reduce
import numpy as np

import textdistance
from openpyxl import workbook, load_workbook

from tqdm import tqdm


class DedupFile:

    def __init__(self, file_name, ignore_same_source=False):
        print("Opening workbook %s" % file_name)
        wb = load_workbook(file_name)
        self.ws = wb.active
        self.ignore_same_source = ignore_same_source
        # Copy entire workbook into memory as an easier-to-access data structure

        # Assume the first column is the ID, which will then be used to identify

        self.ids = list(self.iter_row(1))
        self.source = list(self.iter_row(2))

        self.score = np.ones((len(self.ids), len(self.ids)), np.float16)

        for i in range(len(self.ids)):
            self.score[i, i] = 0

        if ignore_same_source:
            for i in range(len(self.ids)):
                for j in range(len(self.ids)):
                    if i <= j:
                        break
                    if self.source[i] == self.source[j]:
                        self.score[i, j] = 0
                        self.score[j, i] = 0

        self.deduplicate_UEN(3)
        self.deduplicate_column(4)
        self.deduplicate_column(5)
        self.deduplicate_column(6, weight=0.5, dupes_expected=True)

        total_score = sum(np.max(self.score, axis=0))

        print('Total score  is {} for {} rows (average {})'.format(
            total_score, len(self.ids), total_score/len(self.ids)))

        score_column = self.ws.max_column + 1

        self.ws.cell(1, score_column).value = 'Similarity Score'

        for i, max_score in enumerate(np.max(self.score, axis=0)):
            self.ws.cell(i+2, score_column).value = max_score

        wb.save(filename='output.xlsx')

    def deduplicate_column(self, column_no, weight=1, dupes_expected=False):
        '''Returns an n by n table of weights'''
        column_name = self.ws.cell(1, column_no).value
        values = list(self.iter_row(column_no))

        for i, value in tqdm(enumerate(values), total=len(self.ids), desc='Processing {}'.format(column_name)):
            for j, other in enumerate(values):
                if i <= j:
                    break

                if self.ignore_same_source and self.source[i] == self.source[j]:
                    break

                self.compare_values(
                    i, value, j, other, column_no, column_name, weight, (not dupes_expected))

        return

    def deduplicate_UEN(self, column_no):
        # column_name = self.ws.cell(1, column_no).value

        def strip_value(value):
            # strip symbols and spaces
            stripped_value = re.sub('\W', '', value)
            # strip all letter prefixes
            stripped_value = re.sub('^[a-zA-Z]+', '', stripped_value)
            stripped_value = re.sub(
                '^[09]+', '', stripped_value)  # strip leading 0s and 9s
            stripped_value = re.sub(
                '[0]{2,}$', '', stripped_value)  # strip trailing 0s
            return stripped_value

        values = [(value, strip_value(value))
                  for value in self.iter_row(column_no)]

        for i, (value, stripped_value) in tqdm(enumerate(values), total=len(self.ids), desc='Processing UEN'):
            for j, (other, stripped_other) in enumerate(values):
                if i <= j:
                    break

                if self.ignore_same_source and self.source[i] == self.source[j]:
                    break

                if self.is_not_none(value) and self.is_not_none(other):
                    if len(stripped_value) > 0 and len(stripped_other) > 0:
                        dl_striped_similarity = textdistance.levenshtein.normalized_similarity(
                            stripped_value, stripped_other)
                        scaling_factor = (
                            len(stripped_value) + len(stripped_other)) / 5
                    else:
                        dl_striped_similarity = textdistance.levenshtein.normalized_similarity(
                            value, other)
                        scaling_factor = (
                            len(value) + len(other)) / 5

                    if value == other:
                        scaling_factor *= 2

                    self.add_score(i, column_no, j,
                                   (dl_striped_similarity*2) ** scaling_factor)

    def add_score(self, row, col, other_row, score):
        self.score[row, other_row] *= score
        self.score[other_row, row] *= score

    def iter_row(self, column_no):
        return (str(cell[0]).upper() for cell in self.ws.iter_rows(min_row=2, min_col=column_no, max_col=column_no, values_only=True))

    def compare_values(self, i, value, j, other, column_no, column_name, weight=1, bonus_to_same=True):
        '''Compares two values and returns the similarity score. Also logs messages.'''
        if self.is_not_none(value) and self.is_not_none(other):

            scaling_factor = weight*(len(value) + len(other))/20

            if value == other and bonus_to_same:
                scaling_factor *= 1.5

            dl_similarity = textdistance.levenshtein.normalized_similarity(
                value, other)

            # zl_similarity = textdistance.zlib_ncd.normalized_similarity(value, other)
            # score = (max(dl_similarity, zl_similarity)*2) ** scaling_factor
            score = (dl_similarity*2) ** scaling_factor

            self.add_score(i, column_no, j, score)

    @staticmethod
    def is_not_none(value):
        null_values = set(["", "None", "#N/A", "NA", " ", "  ", "-"])
        return value not in null_values


file_name = sys.argv[1]

ignore_same_source = False if len(sys.argv) <= 2 else sys.argv[2]

dedup = DedupFile(file_name, ignore_same_source)
